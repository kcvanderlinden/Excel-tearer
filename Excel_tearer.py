import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import openpyxl

class ExcelFilterAndSave:
    def __init__(self, excel_file_location, column1, column2, *args, **kwargs):
        self.excel_file_location = excel_file_location
        self.column1 = column1
        self.column2 = column2
        self.args = args
        self.kwargs = kwargs
        
        # Ensure the output directory exists
        if not os.path.exists('output'):
            os.makedirs('output')
        
        # Read the Excel file with provided args and kwargs
        self.df = pd.read_excel(self.excel_file_location, *self.args, **self.kwargs)
    
    def filter_and_save(self):
        # Get unique combinations of column1 and column2
        unique_combinations = self.df[[self.column1, self.column2]].drop_duplicates().values.tolist()
        
        for combination in unique_combinations:
            value1, value2 = combination
            
            # Filter the DataFrame based on the current combination
            filtered_df = self.df[(self.df[self.column1] == value1) & (self.df[self.column2] == value2)]
            
            # Create a filename based on the combination values
            file_name = f"{value1}|{value2}.xlsx" # .replace(" ", "_").replace("/", "_")
            output_path = os.path.join('output', file_name)
            
            # Save the filtered DataFrame to an Excel file (without index)
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                filtered_df.to_excel(writer, sheet_name='Sheet1', index=False)
                
                # Load the workbook and select the active worksheet
                workbook = writer.book
                worksheet = writer.sheets['Sheet1']
                
                # Add a table to the worksheet
                table = self.create_table(filtered_df, worksheet)
                worksheet.add_table(table)
                
                # Color the headers in blue and make the text white and bold
                header_fill = PatternFill(start_color='2d5e7f', end_color='2d5e7f', fill_type='solid')
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = openpyxl.styles.Font(color='FFFFFF', bold=True)
                # Auto-adjust column widths
                for column_cells in worksheet.columns:
                    max_length = 0
                    column = column_cells[0].column_letter
                    for cell in column_cells:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column].width = adjusted_width
                # Save the workbook
                workbook.save(output_path)
            
            print(f"Saved subset with {self.column1}={value1}, {self.column2}={value2} to {file_name}")
    
    def create_table(self, df, worksheet):
        table_range = f"A1:{openpyxl.utils.get_column_letter(len(df.columns))}{len(df) + 1}"
        return openpyxl.worksheet.table.Table(displayName="Table", ref=table_range)