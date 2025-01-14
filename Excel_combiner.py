import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import openpyxl

def concatenate_excel_files(directory):
    all_files = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith('.xlsx'):
                all_files.append(os.path.join(root, file))
    print(f"Found {len(all_files)} Excel files in {directory}")
    dfs = []
    for file in all_files:
        df = pd.read_excel(file)
        dfs.append(df)
    return pd.concat(dfs, ignore_index=True)

def report_duplicates(df, column):
    duplicates = df[df.duplicated(subset=column, keep=False)]
    if not duplicates.empty:
        print(f"Duplicate entries in {column}:\n{duplicates}")
    else:
        print(f"No duplicate entries found in {column}.")

def rename_last_columns(df, num_columns, prefix):
    columns_to_rename = df.columns[-num_columns:]
    new_column_names = [f"{prefix} {col}" for col in columns_to_rename]
    df.rename(columns=dict(zip(columns_to_rename, new_column_names)), inplace=True)
    return df

def remove_common_columns(df1, df2, merge_column):
    common_columns = set(df1.columns) & set(df2.columns) - {merge_column}
    df2.drop(columns=common_columns, inplace=True)

def merge_dataframes(df1, df2, merge_column, output_filename):
    merged_df = pd.merge(df1, df2, on=merge_column, how='outer')
    return merged_df

def save_formatted(output_path, merged_df):
    # Save the filtered DataFrame to an Excel file (without index)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        merged_df.to_excel(writer, sheet_name='Sheet1', index=False)
        
        # Load the workbook and select the active worksheet
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # Add a table to the worksheet
        table = create_table(merged_df, worksheet)
        worksheet.add_table(table)
        
        # Color the headers in blue and make the text white and bold
        header_fill = PatternFill(start_color='2d5e7f', end_color='2d5e7f', fill_type='solid')
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = openpyxl.styles.Font(color='FFFFFF', bold=True)
        # Auto-adjust column widths
        for column_cells in worksheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width
        # Save the workbook
        workbook.save(output_path)

def create_table(df, worksheet):
    table_range = f"A1:{openpyxl.utils.get_column_letter(len(df.columns))}{len(df) + 1}"
    return openpyxl.worksheet.table.Table(displayName="Table", ref=table_range)

def process_directories(dir1, dir2, merge_column, last_x_columns, output_filename):
    # Concatenate Excel files from both directories
    df1 = concatenate_excel_files(dir1)
    df2 = concatenate_excel_files(dir2)

    # Column rename based on last part of directory name
    dir1_name = os.path.basename(dir1)
    dir2_name = os.path.basename(dir2)

    # Report on duplicate entries in the specified column
    report_duplicates(df1, merge_column)
    report_duplicates(df2, merge_column)

    # Rename last x amount of columns in both dataframes
    df1 = rename_last_columns(df1, last_x_columns, dir1_name)
    df2 = rename_last_columns(df2, last_x_columns, dir2_name)

    # Remove common columns except the merge column
    remove_common_columns(df1, df2, merge_column)

    # Merge dataframes on the specified column and save to output file
    merged_df = merge_dataframes(df1, df2, merge_column, output_filename)

    # Save the merged DataFrame to an Excel file as a formatted table
    save_formatted(output_filename, merged_df)